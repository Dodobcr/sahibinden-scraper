#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Sahibinden.com Scraper — GitHub Actions versiyonu
Cikti: ilanlar.csv, ilanlar.xlsx, seen.json (repo'ya commit edilir)
"""

import sys, io, json, time, random, os, re, csv, traceback
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from DrissionPage import ChromiumPage, ChromiumOptions
from bs4 import BeautifulSoup

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", line_buffering=True)
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", line_buffering=True)

# ─── CONFIG ───────────────────────────────────────────────────
SAH_URLS      = os.environ.get("SAH_URLS", "")
MAX_LISTINGS  = int(os.environ.get("MAX_LISTINGS", "20"))  # çalışma başına max ilan
MAX_PAGES     = int(os.environ.get("MAX_PAGES", "50"))     # URL başına max sayfa

# Dosyalar repo kökünde tutulur
BASE_DIR  = Path(os.environ.get("GITHUB_WORKSPACE", "."))
CSV_FILE  = BASE_DIR / "ilanlar.csv"
XLSX_FILE = BASE_DIR / "ilanlar.xlsx"
SEEN_FILE = BASE_DIR / "seen.json"

HEADERS = [
    "İlan No", "Link", "İSİM SOYİSİM", "Cep Telefonu",
    "Ada/Parsel", "Kaks (Emsal)", "Emlak Tipi", "İmar Durumu",
    "Tapu Durumu", "m²", "Fiyat", "m² Fiyatı", "İlan Tarihi",
    "Gabari", "Krediye Uygunluk", "Takas",
    "İl", "İlçe", "Mahalle", "Açıklama", "Kimden", "Emlak İsmi",
]

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
]


# ─── SEEN IDs ─────────────────────────────────────────────────

def load_seen() -> set:
    try:
        if SEEN_FILE.exists():
            with open(SEEN_FILE, "r", encoding="utf-8") as f:
                return set(json.load(f))
    except Exception:
        pass
    return set()


def save_seen(seen: set):
    with open(SEEN_FILE, "w", encoding="utf-8") as f:
        json.dump(sorted(list(seen)), f)


# ─── CSV + XLSX ───────────────────────────────────────────────

def init_csv():
    if not CSV_FILE.exists():
        with open(CSV_FILE, "w", newline="", encoding="utf-8-sig") as f:
            csv.writer(f).writerow(HEADERS)


def append_csv(row: dict):
    with open(CSV_FILE, "a", newline="", encoding="utf-8-sig") as f:
        csv.writer(f).writerow([row.get(h, "") for h in HEADERS])


def csv_to_xlsx():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sahibinden İlanlar"
    with open(CSV_FILE, "r", encoding="utf-8-sig") as f:
        for r_idx, row in enumerate(csv.reader(f), 1):
            ws.append(row)
            if r_idx == 1:
                for cell in ws[1]:
                    cell.font      = Font(bold=True, color="FFFFFF")
                    cell.fill      = PatternFill("solid", fgColor="0D5CA8")
                    cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    wb.save(str(XLSX_FILE))
    row_count = ws.max_row - 1
    print(f"[XLSX] Oluşturuldu: {row_count} ilan")
    return row_count


# ─── PARSE ────────────────────────────────────────────────────

def extract_id(url: str) -> str:
    m = re.search(r"-(\d{7,12})/detay", url)
    if m: return m.group(1)
    m = re.search(r"/(\d{7,12})/detay", url)
    if m: return m.group(1)
    m = re.search(r"[/_-](\d{7,12})(?:[/?#]|$)", url)
    if m: return m.group(1)
    return ""


def get_links_from_page(soup) -> list:
    links, seen = [], set()

    # Yontem 1: classifiedTitle class
    for a in soup.find_all("a", class_="classifiedTitle"):
        href = a.get("href", "")
        if href and "/ilan/" in href:
            full = "https://www.sahibinden.com" + href if href.startswith("/") else href
            lid  = extract_id(full)
            if lid and lid not in seen:
                links.append(full)
                seen.add(lid)

    # Yontem 2: /ilan/ iceren tum linkler (classifiedTitle bulunamazsa)
    if not links:
        print("  [DEBUG] classifiedTitle bulunamadi, /ilan/ linkleri deneniyor...", flush=True)
        for a in soup.find_all("a", href=True):
            href = a["href"]
            if "/ilan/" in href and "/detay" in href:
                full = "https://www.sahibinden.com" + href if href.startswith("/") else href
                lid  = extract_id(full)
                if lid and lid not in seen:
                    links.append(full)
                    seen.add(lid)

    # Debug: 0 ilan bulununca sayfa basligini goster
    if not links:
        title = soup.find("title")
        print(f"  [DEBUG] Sayfa title: {title.get_text() if title else 'yok'}", flush=True)
        # Sayfadaki ilk 800 karakteri goster
        text = soup.get_text()[:800].replace('\n', ' ').strip()
        print(f"  [DEBUG] Sayfa icerigi: {text}", flush=True)

    return links


def human_delay():
    """İnsan gibi düzensiz bekleme: bazen hızlı, bazen yavaş, bazen çok yavaş."""
    r = random.random()
    if r < 0.55:
        # %55 — normal okuma (3-10 sn)
        t = random.uniform(3, 10)
    elif r < 0.80:
        # %25 — dikkati dağıldı, biraz bekledi (12-25 sn)
        t = random.uniform(12, 25)
    elif r < 0.93:
        # %13 — telefona baktı (30-55 sn)
        t = random.uniform(30, 55)
    else:
        # %7  — uzun ara (60-120 sn)
        t = random.uniform(60, 120)
    print(f"  [{t:.1f}s bekleniyor]", flush=True)
    time.sleep(t)


def human_scroll(driver):
    try:
        driver.run_js(f"window.scrollBy({{top:{random.randint(300,600)},behavior:'smooth'}})")
        time.sleep(random.uniform(0.8, 1.5))
        driver.run_js(f"window.scrollBy({{top:{random.randint(200,400)},behavior:'smooth'}})")
        time.sleep(random.uniform(0.5, 1.0))
    except Exception:
        pass


def get_phone(driver) -> str:
    try:
        btn = driver.ele("css:.phoneRevealButton", timeout=4)
        if btn:
            time.sleep(random.uniform(0.5, 1.0))
            btn.click()
            time.sleep(random.uniform(2.0, 3.0))
    except Exception:
        pass
    try:
        span = driver.ele("css:#phoneInfoPart .pretty-phone-part span[data-content]", timeout=4)
        if span:
            val = span.attr("data-content")
            if val and val.strip():
                return val.strip()
    except Exception:
        pass
    soup = BeautifulSoup(driver.html, "lxml")
    ul = soup.find("ul", id="phoneInfoPart")
    if ul:
        inner = ul.find("span", attrs={"data-content": True})
        if inner:
            return inner["data-content"].strip()
    return ""


def parse_detail(driver, url: str) -> dict:
    soup = BeautifulSoup(driver.html, "lxml")
    row  = {h: "" for h in HEADERS}
    row["Link"] = url

    seller = ""
    agent_div = soup.find("div", class_="user-info-agent")
    if agent_div:
        h3 = agent_div.find("h3")
        if h3:
            seller = h3.get_text(strip=True)
    if not seller:
        uname_div = soup.find("div", class_="username-info-area")
        if uname_div:
            style = uname_div.find("style")
            if style and style.string and "content:" in style.string:
                m = re.search(r"content:\s*['\"](.+?)['\"]", style.string)
                if m:
                    seller = m.group(1).strip()
    row["İSİM SOYİSİM"] = seller

    store_div = soup.find("div", class_="user-info-store-name")
    if store_div:
        a = store_div.find("a")
        if a:
            row["Emlak İsmi"] = a.get_text(strip=True)

    price_span = soup.find("span", class_="classified-price-wrapper")
    if price_span:
        row["Fiyat"] = price_span.get_text(strip=True)

    info_div = soup.find("div", class_="classifiedInfo")
    if info_div:
        h2 = info_div.find("h2")
        if h2:
            locs = h2.find_all("a")
            if len(locs) > 0: row["İl"]      = locs[0].get_text(strip=True)
            if len(locs) > 1: row["İlçe"]    = locs[1].get_text(strip=True)
            if len(locs) > 2: row["Mahalle"] = locs[2].get_text(strip=True)

    desc_div = soup.find("div", id="classifiedDescription")
    if desc_div:
        row["Açıklama"] = desc_div.get_text(strip=True)[:1000]

    ada = ""
    cl_ul = soup.find("ul", class_="classifiedInfoList")
    if cl_ul:
        for li in cl_ul.find_all("li"):
            strong = li.find("strong")
            span   = li.find("span")
            if not (strong and span):
                continue
            key = strong.get_text(strip=True)
            val = span.get_text(strip=True)
            if   key == "Ada No":            ada = val
            elif key == "Parsel No":         row["Ada/Parsel"]       = f"{ada}/{val}"
            elif key == "Kaks (Emsal)":      row["Kaks (Emsal)"]     = val
            elif key == "Emlak Tipi":        row["Emlak Tipi"]       = val
            elif key == "İmar Durumu":       row["İmar Durumu"]      = val
            elif key == "Tapu Durumu":       row["Tapu Durumu"]      = val
            elif key == "m²":               row["m²"]               = val.replace(".", "")
            elif key == "m² Fiyatı":         row["m² Fiyatı"]        = val.replace(".", "")
            elif key == "İlan Tarihi":       row["İlan Tarihi"]      = val
            elif key == "Gabari":            row["Gabari"]           = val
            elif key == "Krediye Uygunluk":  row["Krediye Uygunluk"] = val
            elif key == "Kimden":            row["Kimden"]           = val
            elif key == "İlan No":           row["İlan No"]          = val
            elif key == "Takas":             row["Takas"]            = val

    human_scroll(driver)
    row["Cep Telefonu"] = get_phone(driver)
    return row


# ─── ANA SCRAPE ───────────────────────────────────────────────

def run_scrape(driver, urls: list, seen: set) -> tuple:
    total_new = total_skip = total_fail = 0

    for u_idx, base_url in enumerate(urls):

        # Limit doldu mu?
        if total_new >= MAX_LISTINGS:
            print(f"\nLimit doldu ({MAX_LISTINGS} ilan). Sonraki çalışmada devam edilecek.", flush=True)
            break

        print(f"\n[{u_idx+1}/{len(urls)}] URL: {base_url}", flush=True)
        driver.get(base_url)
        time.sleep(random.uniform(4, 8))
        page_num = 1

        while page_num <= MAX_PAGES:

            if total_new >= MAX_LISTINGS:
                print(f"  Limit doldu, duruyorum.", flush=True)
                break

            print(f"  Sayfa {page_num}: {driver.url}", flush=True)
            soup  = BeautifulSoup(driver.html, "lxml")
            links = get_links_from_page(soup)
            print(f"  {len(links)} ilan bulundu", flush=True)

            if not links:
                print("  İlan yok, URL tamamlandı.", flush=True)
                break

            for i, det_url in enumerate(links):

                if total_new >= MAX_LISTINGS:
                    break

                ilan_id = extract_id(det_url)
                if not ilan_id:
                    continue
                if ilan_id in seen:
                    total_skip += 1
                    print(f"  ATLANDI #{ilan_id} ({i+1}/{len(links)})", flush=True)
                    continue

                # İlan detayına geçmeden önce insan gibi bekle
                human_delay()

                try:
                    driver.get(det_url)
                    time.sleep(random.uniform(3, 6))
                    row = parse_detail(driver, det_url)
                    if not row.get("İlan No"):
                        row["İlan No"] = ilan_id
                    append_csv(row)
                    seen.add(ilan_id)
                    save_seen(seen)
                    total_new += 1
                    print(f"  ✓ [{total_new}/{MAX_LISTINGS}] #{row.get('İlan No',ilan_id)} | "
                          f"{row.get('Fiyat','?')} | "
                          f"{row.get('İlçe','?')} | "
                          f"tel:{row.get('Cep Telefonu','')}", flush=True)
                except Exception as e:
                    total_fail += 1
                    print(f"  ✗ #{ilan_id}: {str(e)[:150]}", flush=True)

            if total_new >= MAX_LISTINGS:
                break

            # Sonraki sayfa
            next_link = None
            for a in soup.find_all("a", href=True):
                if "Sonraki" in a.get("title", "") or "next" in a.get("rel", []):
                    next_link = a["href"]
                    break
            if not next_link:
                print("  Son sayfa.", flush=True)
                break

            next_url = ("https://www.sahibinden.com" + next_link
                        if next_link.startswith("/") else next_link)
            # Sayfa geçişinde de insan gibi bekle
            pw = random.uniform(15, 35)
            print(f"  Sonraki sayfa – {pw:.1f}s...", flush=True)
            time.sleep(pw)
            driver.get(next_url)
            time.sleep(random.uniform(4, 7))
            page_num += 1

        if u_idx < len(urls) - 1 and total_new < MAX_LISTINGS:
            w = random.randint(20, 40)
            print(f"\nSonraki URL – {w}s bekleniyor...", flush=True)
            time.sleep(w)

    return total_new, total_skip, total_fail


# ─── MAIN ─────────────────────────────────────────────────────

def main():
    print("=" * 60, flush=True)
    print(f"Sahibinden Scraper — {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", flush=True)
    print("=" * 60, flush=True)

    urls = [u.strip() for u in SAH_URLS.splitlines() if u.strip()]
    if not urls:
        print("HATA: SAH_URLS boş! GitHub Secrets'ı kontrol edin.")
        sys.exit(1)
    print(f"URL sayısı: {len(urls)}", flush=True)

    init_csv()
    seen = load_seen()
    print(f"Daha önce çekilmiş: {len(seen)} ilan", flush=True)

    # Chrome başlat
    print("\n[CHROME] Başlatılıyor...", flush=True)
    opts = ChromiumOptions()
    opts.set_argument("--no-sandbox")
    opts.set_argument("--disable-dev-shm-usage")
    opts.set_argument("--disable-blink-features=AutomationControlled")
    opts.set_argument(f"--user-agent={random.choice(USER_AGENTS)}")
    opts.set_argument("--lang=tr-TR,tr;q=0.9")
    opts.set_argument("--disable-notifications")
    opts.set_argument("--window-size=1920,1080")
    opts.set_argument("--disable-gpu")
    # DISPLAY varsa xvfb kullan (headless degil - Cloudflare bypass)
    if not os.environ.get("DISPLAY"):
        opts.headless()
        print("[CHROME] Headless mod (DISPLAY yok)", flush=True)
    else:
        print(f"[CHROME] Sanal ekran modu DISPLAY={os.environ.get('DISPLAY')}", flush=True)

    driver = ChromiumPage(addr_or_opts=opts)
    try:
        new_cnt, skip_cnt, fail_cnt = run_scrape(driver, urls, seen)
    except Exception:
        print(f"[HATA]\n{traceback.format_exc()}", flush=True)
        new_cnt = skip_cnt = fail_cnt = 0
    finally:
        try:
            driver.quit()
        except Exception:
            pass

    print(f"\n{'='*60}", flush=True)
    print(f"TAMAMLANDI: {new_cnt} yeni | {skip_cnt} atlandı | {fail_cnt} hata", flush=True)
    print(f"{'='*60}", flush=True)

    # Excel oluştur
    csv_to_xlsx()

    # Toplam ilan sayısını env'e yaz (commit mesajı için)
    if os.environ.get("GITHUB_ENV"):
        with open(os.environ["GITHUB_ENV"], "a") as f:
            f.write(f"NEW_COUNT={new_cnt}\n")


if __name__ == "__main__":
    main()
